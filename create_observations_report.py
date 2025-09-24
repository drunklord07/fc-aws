import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pandas.api.types import CategoricalDtype
from tqdm import tqdm

def create_observations_sheet_report():
    current_directory = os.path.dirname(os.path.realpath(__file__))
    csv_file = None
    for file in os.listdir(current_directory):
        if file.endswith('.csv'):
            csv_file = file
            break
    
    if not csv_file:
        print("Error: No CSV file found in the current folder for 'create_observations_report.py'.")
        return

    input_csv_path = os.path.join(current_directory, csv_file)
    output_excel_path = os.path.join(current_directory, 'report.xlsx')

    try:
        print("--- Starting 'Observations' Report Generation ---")
        print(f"1. Reading CSV file: {csv_file}")
        df = pd.read_csv(input_csv_path, sep=';')

        print("2. Deleting rows with severity 'INFO'.")
        df = df[df['SEVERITY'].str.upper() != 'INFO']

        # Drop unnecessary columns
        columns_to_delete = [
            'ASSESSMENT_START_TIME', 'FINDING_UNIQUE_ID', 'PROVIDER', 'CHECK_ID', 
            'CHECK_TYPE', 'SUBSERVICE_NAME', 'RESOURCE_TYPE', 'RESOURCE_DETAILS', 
            'RESOURCE_TAGS', 'RELATED_URL', 'REMEDIATION_RECOMMENDATION_CODE_NATIVEIAC', 
            'REMEDIATION_RECOMMENDATION_CODE_TERRAFORM', 'REMEDIATION_RECOMMENDATION_CODE_CLI', 
            'REMEDIATION_RECOMMENDATION_CODE_OTHER', 'COMPLIANCE', 'CATEGORIES', 'DEPENDS_ON', 
            'RELATED_TO', 'NOTES', 'PROFILE', 'ACCOUNT_ID', 'ACCOUNT_NAME', 'ACCOUNT_EMAIL', 
            'ACCOUNT_ARN', 'ACCOUNT_ORG', 'ACCOUNT_TAGS'
        ]
        
        print("3. Deleting specified columns from the DataFrame.")
        df.drop(columns=[c for c in columns_to_delete if c in df.columns], inplace=True)

        # Handle variations in resource column naming
        if 'RESOURCE_ARN' not in df.columns:
            for alt_col in ['RESOURCE_UID', 'RESOURCE_ID']:
                if alt_col in df.columns:
                    df.rename(columns={alt_col: 'RESOURCE_ARN'}, inplace=True)
                    break

        # Normalize severity
        severity_order = ['Critical', 'High', 'Medium', 'Low']
        if 'SEVERITY' in df.columns:
            df['SEVERITY'] = df['SEVERITY'].str.capitalize()
            severity_cat_type = CategoricalDtype(categories=severity_order, ordered=True)
            df['SEVERITY'] = df['SEVERITY'].astype(severity_cat_type)

        # Normalize STATUS (only PASS/FAIL mapped, rest untouched)
        if 'STATUS' in df.columns:
            df['STATUS'] = df['STATUS'].astype(str).str.strip().str.upper().replace({
                'PASS': 'Compliant',
                'FAIL': 'Non-Compliant'
            })

        print("4. Sorting DataFrame by Severity and Control (CHECK_TITLE).")
        sort_columns = []
        if 'SEVERITY' in df.columns:
            sort_columns.append('SEVERITY')
        if 'CHECK_TITLE' in df.columns:
            sort_columns.append('CHECK_TITLE')
        
        if sort_columns:
            df = df.sort_values(by=sort_columns, ascending=[True, True])

        print("5. Adding 'S. No' column.")
        df.insert(0, 'S. No', None) 

        desired_column_order = [
            'S. No', 'CHECK_TITLE', 'SEVERITY', 'REGION', 'SERVICE_NAME', 'RESOURCE_ARN', 
            'STATUS', 'STATUS_EXTENDED', 'DESCRIPTION', 'RISK', 
            'REMEDIATION_RECOMMENDATION_TEXT', 'REMEDIATION_RECOMMENDATION_URL'
        ]

        print("6. Reordering columns.")
        final_columns = [col for col in desired_column_order if col in df.columns]
        df = df[final_columns]

        print(f"7. Writing DataFrame to Excel file: {output_excel_path}")
        df.to_excel(output_excel_path, index=False, engine='openpyxl')

        print("8. Loading Excel file for styling 'Observations' sheet.")
        wb = load_workbook(output_excel_path)
        ws = wb.active
        ws.title = "Observations"

        header_rename_map = {
            'S. No': 'S. No', 'CHECK_TITLE': 'Control', 'SEVERITY': 'Severity', 'REGION': 'Region', 
            'SERVICE_NAME': 'Service', 'RESOURCE_ARN': 'Resource ARN', 'STATUS': 'Status', 
            'STATUS_EXTENDED': 'Observation', 'DESCRIPTION': 'Description', 'RISK': 'Impact', 
            'REMEDIATION_RECOMMENDATION_TEXT': 'Remediation', 'REMEDIATION_RECOMMENDATION_URL': 'Reference'
        }

        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        header_fill = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")
        header_font = Font(name='Arial', size=10, bold=True, color="FFFFFF")

        print("9. Applying header styles and renaming to 'Observations' sheet.")
        for col_idx, cell in enumerate(ws[1]):
            original_header = cell.value
            if original_header in header_rename_map:
                cell.value = header_rename_map[original_header]
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.fill = header_fill
            cell.border = thin_border

        # Column index lookup
        headers = [cell.value for cell in ws[1]]
        s_no_col_idx = headers.index('S. No')+1 if 'S. No' in headers else -1
        severity_col_idx = headers.index('Severity')+1 if 'Severity' in headers else -1
        status_col_idx = headers.index('Status')+1 if 'Status' in headers else -1

        # Apply numbering to S. No column
        if s_no_col_idx != -1:
            for row_idx in range(2, ws.max_row + 1):
                s_no_cell = ws.cell(row=row_idx, column=s_no_col_idx)
                s_no_cell.value = f"=ROW()-1"
                s_no_cell.alignment = Alignment(horizontal='center', vertical='center')
                s_no_cell.border = thin_border

        print("10. Setting column widths for key columns.")
        column_widths = {
            s_no_col_idx: 7, severity_col_idx: 13, status_col_idx: 14
        }
        for col_idx, width in column_widths.items():
            if col_idx != -1:
                ws.column_dimensions[get_column_letter(col_idx)].width = width

        severity_colors = {
            'Low': PatternFill(start_color="92d050", end_color="92d050", fill_type="solid"), 
            'Medium': PatternFill(start_color="ffc000", end_color="ffc000", fill_type="solid"),
            'High': PatternFill(start_color="ff0000", end_color="ff0000", fill_type="solid"), 
            'Critical': PatternFill(start_color="c00000", end_color="c00000", fill_type="solid")
        }
        status_colors = {
            'Compliant': PatternFill(start_color="92d050", end_color="92d050", fill_type="solid"),
            'Non-Compliant': PatternFill(start_color="ff0000", end_color="ff0000", fill_type="solid")
        }

        print("11. Applying coloring only to Severity and Status columns.")
        for row_idx in tqdm(range(2, ws.max_row + 1), desc="Formatting Severity/Status"):
            if severity_col_idx != -1:
                severity_cell = ws.cell(row=row_idx, column=severity_col_idx)
                sev_value = str(severity_cell.value).capitalize()
                if sev_value in severity_colors:
                    severity_cell.fill = severity_colors[sev_value]
                    if sev_value == 'Critical':
                        severity_cell.font = Font(name='Arial', size=10, bold=True, color="FFFFFF")
                    else:
                        severity_cell.font = Font(name='Arial', size=10, bold=True)
                severity_cell.alignment = Alignment(horizontal='center', vertical='center')
            if status_col_idx != -1:
                status_cell = ws.cell(row=row_idx, column=status_col_idx)
                status_value = str(status_cell.value).strip()
                if status_value in status_colors:
                    status_cell.fill = status_colors[status_value]
                    status_cell.font = Font(name='Arial', size=10, bold=True)
                status_cell.alignment = Alignment(horizontal='center', vertical='center')

        print("12. Disabling gridlines for 'Observations' sheet.")
        ws.sheet_view.showGridlines = False

        print(f"13. Saving Excel file with 'Observations' sheet: {output_excel_path}")
        wb.save(output_excel_path)
        print(f"'Observations' report generation complete! File saved at: {output_excel_path}")
        
    except Exception as e:
        print(f"An error occurred during 'Observations' report generation: {e}")

if __name__ == "__main__":
    create_observations_sheet_report()
