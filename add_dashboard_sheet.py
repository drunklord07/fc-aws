import os
import re
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import column_index_from_string
from datetime import date
from openpyxl.chart import PieChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.chart.label import DataLabelList
from openpyxl.drawing.fill import ColorChoice
from openpyxl.drawing.line import LineProperties
from openpyxl.chart.shapes import GraphicalProperties

def add_dashboard_sheet_and_image():
    current_directory = os.path.dirname(os.path.realpath(__file__))
    excel_file_path = os.path.join(current_directory, 'report.xlsx')
    image_path = os.path.join(current_directory, 'GT.png')

    if not os.path.exists(excel_file_path):
        print(f"Error: Excel file '{excel_file_path}' not found. Please run 'create_observations_report.py' first.")
        return

    # Extract AWS Account ID
    aws_account_id = "N/A_Account_ID"
    prowler_csv_pattern = re.compile(r'prowler-output-(\d+)-\d+\.csv')

    for file_name in os.listdir(current_directory):
        match = prowler_csv_pattern.match(file_name)
        if match:
            aws_account_id = match.group(1)
            print(f"Extracted AWS Account ID: {aws_account_id}")
            break

    try:
        print("--- Starting Dashboard Creation ---")
        wb = load_workbook(excel_file_path)

        # Count severities from Observations sheet
        ws_observations = wb['Observations']
        severity_counts = {
            'Critical': 0,
            'High': 0,
            'Medium': 0,
            'Low': 0,
            'Informational': 0
        }

        # Dynamically find the Severity column
        severity_col_idx = None
        for col_idx, cell in enumerate(ws_observations[1], 1):
            if cell.value == 'Severity':
                severity_col_idx = col_idx
                break

        if severity_col_idx is None:
            print("Error: 'Severity' column not found in 'Observations' sheet.")
            return

        for row in ws_observations.iter_rows(min_row=2, min_col=severity_col_idx, max_col=severity_col_idx):
            severity = str(row[0].value).capitalize()
            if severity in severity_counts:
                severity_counts[severity] += 1

        total_count = sum(severity_counts.values())
        print(f"Severity Counts: {severity_counts}, Total={total_count}")

        # Create/Reset Dashboard sheet
        if 'Dashboard' in wb.sheetnames:
            wb.remove(wb['Dashboard'])
        ws_dashboard = wb.create_sheet(title="Dashboard", index=0)
        ws_dashboard.sheet_view.showGridlines = False

        # --- Create a hidden sheet for chart data ---
        if 'Severity Chart Data' in wb.sheetnames:
            wb.remove(wb['Severity Chart Data'])
        ws_chart_data = wb.create_sheet(title="Severity Chart Data")
        ws_chart_data.sheet_state = 'hidden'

        # Write data to the hidden sheet
        ws_chart_data['A1'] = "Severity"
        ws_chart_data['B1'] = "Count"

        # Ensure consistent order for chart data (Critical, High, Medium, Low)
        chart_severities_order = ['Critical', 'High', 'Medium', 'Low']

        for i, severity in enumerate(chart_severities_order):
            ws_chart_data.cell(row=i+2, column=1, value=severity)
            ws_chart_data.cell(row=i+2, column=2, value=severity_counts.get(severity, 0))


        # Styles
        thick_border = Border(left=Side(style='thick'), right=Side(style='thick'),
                              top=Side(style='thick'), bottom=Side(style='thick'))
        header_purple_fill = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")
        grey_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        white_bold_font = Font(name='Arial', size=12, bold=True, color="FFFFFF")
        black_bold_font = Font(name='Arial', size=10, bold=True, color="000000")
        black_regular_font = Font(name='Arial', size=10, bold=False, color="000000")

        # Add Image
        if os.path.exists(image_path):
            try:
                img = Image(image_path)
                img.height = 66
                img.width = 448
                img.anchor = 'B3'
                ws_dashboard.add_image(img)
                print("Logo image added successfully")
            except Exception as e:
                print(f"Image error: {e}")

        # Add Header Text Box
        text_box_range = 'B7:G12'
        ws_dashboard.merge_cells(text_box_range)
        header_cell = ws_dashboard['B7']
        header_cell.value = "XYX Ltd\n\nAWS Secure Architecture Review\n\nDraft Report\n\n" + date.today().strftime("%d %B %Y")
        header_cell.fill = header_purple_fill
        header_cell.font = white_bold_font
        header_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

        # Add Scope/Duration Table
        ws_dashboard['B15'] = "Scope"
        ws_dashboard['B15'].fill = grey_fill
        ws_dashboard['B15'].font = black_bold_font
        ws_dashboard['B15'].alignment = Alignment(horizontal='center', vertical='center')
        ws_dashboard['B15'].border = thick_border

        ws_dashboard.merge_cells('C15:G15')
        ws_dashboard['C15'] = f"AWS Account - {aws_account_id}"
        ws_dashboard['C15'].fill = white_fill
        ws_dashboard['C15'].font = black_bold_font
        ws_dashboard['C15'].alignment = Alignment(horizontal='center', vertical='center')
        ws_dashboard['C15'].border = thick_border

        ws_dashboard['B16'] = "Duration of Assessment"
        ws_dashboard['B16'].fill = grey_fill
        ws_dashboard['B16'].font = black_bold_font
        ws_dashboard['B16'].alignment = Alignment(horizontal='center', vertical='center')
        ws_dashboard['B16'].border = thick_border

        ws_dashboard.merge_cells('C16:G16')
        ws_dashboard['C16'] = "06.03.2025 - 25.03.2025"
        ws_dashboard['C16'].fill = white_fill
        ws_dashboard['C16'].font = black_regular_font
        ws_dashboard['C16'].alignment = Alignment(horizontal='center', vertical='center')
        ws_dashboard['C16'].border = thick_border

        # Add Summary Table
        severity_colors = {
            'Critical': ("c00000", "FFFFFF"),
            'High': ("ff0000", "000000"),
            'Medium': ("ffc000", "000000"),
            'Low': ("92d050", "000000"),
            'Total': ("D9D9D9", "000000")
        }

        ws_dashboard.merge_cells('J14:N14')
        ws_dashboard['J14'] = "Summary of Observations"
        ws_dashboard['J14'].fill = header_purple_fill
        ws_dashboard['J14'].font = white_bold_font
        ws_dashboard['J14'].alignment = Alignment(horizontal='center', vertical='center')
        ws_dashboard['J14'].border = thick_border

        headers = ["Critical", "High", "Medium", "Low", "Total"]
        for col_idx, header in enumerate(headers, start=10):
            cell = ws_dashboard.cell(row=15, column=col_idx)
            cell.value = header
            cell.fill = PatternFill(start_color=severity_colors[header][0],
                                    end_color=severity_colors[header][0],
                                    fill_type="solid")
            cell.font = Font(name='Arial', size=10, bold=True, color=severity_colors[header][1])
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thick_border

        values = [severity_counts['Critical'], severity_counts['High'],
                  severity_counts['Medium'], severity_counts['Low'], total_count]
        for col_idx, value in enumerate(values, start=10):
            cell = ws_dashboard.cell(row=16, column=col_idx)
            cell.value = value
            cell.fill = white_fill
            cell.font = black_regular_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thick_border

        # ====== PIE CHART - LATEST MODIFICATIONS ======
        pie_chart = PieChart()
        pie_chart.title = "Summary of Observations"
        pie_chart.title.font = Font(name='Arial', size=16, bold=True)

        # Data references pointing to the hidden 'Severity Chart Data' sheet
        labels = Reference(ws_chart_data, min_col=1, min_row=2, max_col=1, max_row=5)
        data = Reference(ws_chart_data, min_col=2, min_row=2, max_col=2, max_row=5)

        pie_chart.add_data(data, titles_from_data=False)
        pie_chart.set_categories(labels)

        # Data labels configuration - RE-ADDED and set to INSIDE END
        pie_chart.dataLabels = DataLabelList()
        pie_chart.dataLabels.showVal = True
        pie_chart.dataLabels.showPercent = True
        pie_chart.dataLabels.font = Font(name='Arial', size=8, bold=True) # Slightly smaller font for very small chart
        pie_chart.dataLabels.dLblPos = 'inEnd' # Position labels INSIDE END

        # Legend position - SET TO BOTTOM
        pie_chart.legend.position = 'b'

        # Configure slice colors (Critical to Low)
        colors = ["c00000", "ff0000", "ffc000", "92d050"]
        series = pie_chart.series[0]
        data_points = []

        for i, color in enumerate(colors):
            dp = DataPoint(idx=i)
            dp.graphicalProperties = GraphicalProperties(
                solidFill=ColorChoice(srgbClr=color)
            )
            data_points.append(dp)

        # Assign data points to series
        series.data_points = data_points

        # Current chart size (from previous version): 5% of 5x2 inches (approx 0.25x0.10 inches)
        # Making it 50% smaller means multiplying by 0.5
        current_width_px = (5 * 96) * 0.05
        current_height_px = (2 * 96) * 0.05

        pie_chart.width = current_width_px * 0.50
        pie_chart.height = current_height_px * 0.50

        # Purple border styling
        pie_chart.graphicalProperties = GraphicalProperties(
            ln=LineProperties(
                solidFill=ColorChoice(srgbClr="7030A0"),
                w=12700   # 1pt border width
            )
        )

        # Position chart at 'J3'
        ws_dashboard.add_chart(pie_chart, 'J3')
        # ====== END PIE CHART ======

        # Adjust column widths and row heights
        ws_dashboard.column_dimensions['B'].width = 25
        for col in ['J', 'K', 'L', 'M', 'N']:
            ws_dashboard.column_dimensions[col].width = 12

        # Save the workbook
        wb.save(excel_file_path)
        print(f"Dashboard created successfully at: {excel_file_path}")
        print(f"Pie chart size: {pie_chart.width/96:.3f} inches wide x {pie_chart.height/96:.3f} inches tall")
        print(f"(Actual pixel size: {pie_chart.width}px wide x {pie_chart.height}px tall)")

    except PermissionError:
        print(f"ERROR: Close Excel file before running script: {excel_file_path}")
    except Exception as e:
        print(f"Unexpected error: {str(e)}")

if __name__ == "__main__":
    add_dashboard_sheet_and_image()